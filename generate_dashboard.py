import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from calculate_dashboard import GenerateReports, GenerateReportsVmd2
import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from config import INPUT_BASE_DIR, OUTPUT_BASE_DIR

class WritingValue:
    def __init__(self, total_attempt_call, total_login_agent, total_success_call, product_counts, status):
        self.monthFormat = datetime.datetime(2025, 5, 10).strftime("%b")
        self.outputFile = f"{OUTPUT_BASE_DIR}\\Daily_Report_{self.monthFormat}.xlsx"
        self.total_attempt_call = total_attempt_call
        self.total_login_agent = total_login_agent
        self.total_success_call = total_success_call
        self.product_counts = product_counts
        self.status = status
        self.datetime = datetime.datetime(2025, 5, 10).strftime("%d-%b")

    def writingValueInExcel(self):
        workbook = load_workbook(self.outputFile)
        sheet = workbook.active
        max_col = sheet.max_column + 1 if self.status == 'team 1' else sheet.max_column
        max_col_letter = get_column_letter(max_col)

        df = self._extract_sheet_data(sheet)
        filtered_df = df[df['cell_value'].notna()].copy()
        filtered_df['cell_coordinate'] = filtered_df['cell_coordinate'].astype(str).str.extract(r'(\d+)$')
        # print(filtered_df)

        indices = self._get_indices(filtered_df)
        summary_df, summary_coords = self._prepare_summary_df(filtered_df, max_col_letter, indices)
        
        # print(summary_df, summary_coords)
        label_df = self._prepare_label_df(filtered_df, max_col_letter)
        team_df = self._prepare_team_df(filtered_df, max_col_letter, indices)
        
        # print("summary coord", summary_coords)

        productOffer = pd.read_csv(f"{INPUT_BASE_DIR}\\Product_Offer\\ProductOffer.csv")
        grouped, offer_names, headers_to_list = self._group_labels(label_df, productOffer)
        agent_summary_df = self._get_agent_summary_df(grouped, offer_names, headers_to_list)
        filtered_summary_df = self._get_filtered_summary_df(agent_summary_df)
        summary_coords.update(self._get_summary_coords(agent_summary_df, filtered_summary_df, summary_df))
    
        merged_product_counts = self._merge_product_counts(productOffer)
        self._fill_team_df(team_df, merged_product_counts)
        self._write_team_df_to_sheet(sheet, team_df)

        product_df = self._prepare_product_df(grouped, productOffer)
        total_revenue_formula = self._build_total_revenue_formula(product_df)
        product_total_formula = self._write_product_formulas(sheet, product_df)
        self._write_summary_formulas(sheet, filtered_summary_df, summary_coords, total_revenue_formula, product_total_formula)
        self._style_sheet(sheet, label_df, self.datetime)
        
        
        workbook.save(self.outputFile)
        
        print("Workbook saved successfully with the updated values.")

    def _extract_sheet_data(self, sheet):
        data = []
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                data.append({'cell_coordinate': cell.coordinate, 'cell_value': cell.value})
        return pd.DataFrame(data)

    def _get_indices(self, df):
        indices = {}
        if self.status == 'team 1':
            indices['start'] = df[df['cell_value'] == 'Team 1'].index.min()
            indices['end'] = df[df['cell_value'] == 'Team 2'].index.min() - 2
        elif self.status == 'team 2':
            indices['start'] = df[df['cell_value'] == 'Team 2'].index.min()
            indices['end'] = df.last_valid_index()
        indices['summary_start'] = 2
        indices['summary_end'] = df[df['cell_value'] == 'Team 1'].index.min()
        return indices

    def _prepare_summary_df(self, df, col_letter, indices):
        summary_df = df.loc[indices['summary_start']:indices['summary_end'] - 1].copy()
        summary_df['cell_coordinate'] = summary_df['cell_coordinate'].apply(lambda x: col_letter + x)
        last_idx = summary_df.last_valid_index()
        summary_coords = {
            'total_summary': summary_df.loc[last_idx, 'cell_coordinate'],
            'pack_sale': summary_df[summary_df['cell_value'] == 'Pack Sale']['cell_coordinate'].iloc[0]
        }
        
        # print(summary_coords)
        return summary_df, summary_coords

    def _prepare_label_df(self, df, col_letter):
        label_df = df.copy()
        label_df['cell_coordinate'] = label_df['cell_coordinate'].apply(lambda x: col_letter + x)
        return label_df

    def _prepare_team_df(self, df, col_letter, indices):
        team_df = df.loc[indices['start']:indices['end']].copy()
        team_df['cell_coordinate'] = team_df['cell_coordinate'].apply(lambda x: col_letter + x)
        team_df['value'] = None
        return team_df

    def _group_labels(self, label_df, productOffer):
        grouped = label_df.dropna(subset=['cell_value']).groupby('cell_value')['cell_coordinate'].apply(list)
        offer_names = productOffer['Offer Name'].tolist()
        headers = ['Team 1', 'Team 2', 'Summary', 'Activities', 'Total', 'Pack Sale']
        return grouped, offer_names, headers

    def _get_agent_summary_df(self, grouped, offer_names, headers):
        df = grouped.reset_index()[['cell_value', 'cell_coordinate']]
        agent_summary_df = df[~df['cell_value'].isin(offer_names + headers)].dropna(subset=['cell_value'])
        return agent_summary_df

    def _get_filtered_summary_df(self, agent_summary_df):
        return agent_summary_df[agent_summary_df['cell_value'].isin([
            'Total Attempts Calls', 'Total Login Agent', 'Total Success Call'
        ])]

    def _get_summary_coords(self, agent_summary_df, filtered_summary_df, summary_df):
        coords = {
            'success_pack': agent_summary_df[agent_summary_df['cell_value'] == '% Success Call Vs Pack Sales']['cell_coordinate'].iloc[0][0],
            'revenue': agent_summary_df[agent_summary_df['cell_value'] == 'Total Revenue']['cell_coordinate'].iloc[0][0],
            'success_cell': agent_summary_df[agent_summary_df['cell_value'] == '% Of Total Vs Success Call']['cell_coordinate'].iloc[0][0],
            'attempt_cell': filtered_summary_df[filtered_summary_df['cell_value'] == 'Total Attempts Calls']['cell_coordinate'].iloc[0][0],
            'success_call': filtered_summary_df[filtered_summary_df['cell_value'] == 'Total Success Call']['cell_coordinate'].iloc[0][0],
        }
        return coords

    def _merge_product_counts(self, productOffer):
        merged = pd.merge(self.product_counts, productOffer, left_on='OFFERID', right_on='Product ID', how='left')
        merged = merged.drop('Product ID', axis=1)
        merged['price_per_pack'] = merged['Offer Name'].str.extract(r'(\d+)\s*ks')[0].astype(int)
        merged['total_amount'] = merged['COUNT OF PACK SALES'] * merged['price_per_pack']
        return merged

    def _fill_team_df(self, team_df, merged_product_counts):
        # Find indices
        idx = lambda label: team_df[team_df['cell_value'] == label].index.min()
        team_df.loc[idx('Total Attempts Calls'), 'value'] = self.total_attempt_call['Total Attempts Calls'][0]
        team_df.loc[idx('Total Login Agent'), 'value'] = self.total_login_agent['Total Login Agent'][0]
        team_df.loc[idx('Total Success Call'), 'value'] = self.total_success_call['Total Success Call'][0]

        attempts = self.total_attempt_call['Total Attempts Calls'][0]
        successes = self.total_success_call['Total Success Call'][0]
        attemptsSuccessPercent = round((successes / attempts) * 100, 0)
        packSaleSuccessPercent = round((merged_product_counts['COUNT OF PACK SALES'].sum() / successes) * 100, 0)

        team_df.loc[idx('% Of Total Vs Success Call'), 'value'] = f"{attemptsSuccessPercent} %"
        team_df.loc[idx('% Success Call Vs Pack Sales'), 'value'] = f"{packSaleSuccessPercent} %"
        team_df.loc[idx('Total Revenue'), 'value'] = merged_product_counts['total_amount'].sum()
        team_df.loc[team_df.last_valid_index(), 'value'] = merged_product_counts['COUNT OF PACK SALES'].sum()
        
        # print(merged_product_counts)

        for i, offer_name in enumerate(merged_product_counts['Offer Name']):
            team_df.loc[team_df['cell_value'] == offer_name, 'value'] = merged_product_counts['COUNT OF PACK SALES'].iloc[i]
            

    def _write_team_df_to_sheet(self, sheet, team_df):
        for _, row in team_df.iterrows():
            if pd.notna(row['value']):
                sheet[row['cell_coordinate']] = row['value']

    def _prepare_product_df(self, grouped, productOffer):
        product_df = pd.merge(grouped, productOffer, left_on='cell_value', right_on='Offer Name', how='left')
        product_df = product_df.dropna(subset=['Product ID', 'Offer Name']).copy()
        product_df['price_per_pack'] = product_df['Offer Name'].str.extract(r'(\d+)\s*ks')[0].astype(int)
        return product_df

    def _build_total_revenue_formula(self, product_df):
        formula = '+'.join([f"({coords[0]} * {price})" for coords, price in zip(product_df['cell_coordinate'], product_df['price_per_pack'])])
        return f"={formula}"

    def _write_product_formulas(self, sheet, product_df):
        total_product_formula = ','.join([coords[0] for coords in product_df['cell_coordinate']])
        for coords in product_df['cell_coordinate']:
            sheet[coords[0]] = f"=SUM({coords[1]},{coords[2]})"
        return f"=SUM({total_product_formula})"

    def _write_summary_formulas(self, sheet, filtered_summary_df, summary_coords, total_revenue_formula, product_total_formula):
        # Write summary formulas
        for cell in filtered_summary_df['cell_coordinate']:
            sheet[cell[0]] = f"=SUM({cell[1]},{cell[2]})"
        
        sheet[summary_coords['success_cell']] = f'=TEXT(ROUND({summary_coords["success_call"]}/{summary_coords["attempt_cell"]}*100, 0), "0") & "%"'
        sheet[summary_coords['success_pack']] = f'=TEXT(ROUND({summary_coords["total_summary"]}/{summary_coords["success_call"]}*100, 0), "0") & "%"'
        sheet[summary_coords['revenue']] = total_revenue_formula
        sheet[summary_coords['total_summary']] = product_total_formula
        
        # print(summary_coords)

    def _style_sheet(self, sheet, label_df, datetime_str):
        fill = PatternFill(start_color='a5d1f2', end_color='a5d1f2', fill_type='solid')
        bold_font = Font(bold=True)
        summary_header_font = Font(bold=True, size=14)
        sheet['A1'].font = summary_header_font
        sheet.freeze_panes = None
        alignment = Alignment(horizontal='right')
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = alignment
        label_cells = label_df[label_df['cell_value'].isin(['Activities', 'Pack Sale'])]['cell_coordinate']
        for cell_coord in label_cells:
            sheet[cell_coord] = datetime_str
            
            sheet[f'A{cell_coord[1:]}'].fill = fill
            sheet[cell_coord].fill = fill
            sheet[f'A{cell_coord[1:]}'].font = bold_font
            sheet[cell_coord].font = bold_font

class WritingValueTeam2(WritingValue):
    pass

def generate_dashboard_main():
    datetime_vrd = datetime.datetime(2025, 5, 10).strftime("%d-%b-%Y")
    
    # Team 1
    reports_vmd1 = GenerateReports(f"{INPUT_BASE_DIR}\\Formatted_VMD\\{datetime_vrd}\\OUTBOUND_REPORT_VMD1.csv")
    total_attempt_call = pd.DataFrame([reports_vmd1.total_attempts_call()], columns=['Total Attempts Calls'])
    total_login_agent = pd.DataFrame([reports_vmd1.total_login_agents()], columns=['Total Login Agent'])
    total_success_call = pd.DataFrame([reports_vmd1.total_success_calls()], columns=['Total Success Call'])
    product_counts = reports_vmd1.product_counts()
    resultTeam1 = WritingValue(total_attempt_call, total_login_agent, total_success_call, product_counts, status="team 1")
    resultTeam1.writingValueInExcel()

    # Team 2
    reports_vmd2 = GenerateReportsVmd2(f"{INPUT_BASE_DIR}\\Formatted_VMD\\{datetime_vrd}\\OUTBOUND_REPORT_VMD2.csv")
    total_attempt_call = pd.DataFrame([reports_vmd2.total_attempts_call()], columns=['Total Attempts Calls'])
    total_login_agent = pd.DataFrame([reports_vmd2.total_login_agents()], columns=['Total Login Agent'])
    total_success_call = pd.DataFrame([reports_vmd2.total_success_calls()], columns=['Total Success Call'])
    product_counts = reports_vmd2.product_counts()
    resultTeam2 = WritingValueTeam2(total_attempt_call, total_login_agent, total_success_call, product_counts, status="team 2")
    resultTeam2.writingValueInExcel()

if __name__ == "__main__":
    generate_dashboard_main()