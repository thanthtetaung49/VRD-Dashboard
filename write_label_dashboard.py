import pandas as pd
from openpyxl import load_workbook, Workbook

class CreateLabel: 
    def __init__(self, outputFile, summary, team1, team2, offerName):
        self.outputFile = outputFile
        self.summary = summary
        self.team1 = team1
        self.team2 = team2
        self.offerName = offerName
        
    def lastRowNumbers(self, sheet):
        lastRows = 0
        for col in sheet.iter_cols(min_col=1):
            cellCoordinate = [cell.coordinate for cell in col]
            numbers = [int(number[1:]) for number in cellCoordinate]
            lastIndex = len(numbers) - 1
            lastRows = numbers[lastIndex]
            
        return lastRows

    def writeLabel(self):
        outputFile = self.outputFile
        summaryLabel = self.summary
        team1Label = self.team1
        team2Label = self.team2
        
        # print(self.offerName.at[0, 'Offer Name'] = 'Pack Sale')
        self.offerName.rename(columns={'Offer Name': 'Pack Sale'}, inplace=True)
        
        new_row = pd.DataFrame({'Pack Sale': ['Pack Sale']})
        self.offerName = pd.concat([new_row, self.offerName], ignore_index=True)
        
        new_row_end = pd.DataFrame({'Pack Sale': ['Total']})
        self.offerName = pd.concat([self.offerName, new_row_end], ignore_index=True)
        
        offerName = self.offerName['Pack Sale']
        
        summaryPackSaleStartPoint = 0
        team1AgentStartPoint, team1PackSaleStartPoint = 0, 0
        team2AgentStartPoint, team2PackSaleStartPoint = 0, 0
        
        try: 
            workbook = load_workbook(outputFile)
            
            sheet = workbook.active
            lastRows = self.lastRowNumbers(sheet)

            # summary agent
            for index,name in enumerate(summaryLabel):
                sheet[f'A{index + 1}'] = name
                summaryPackSaleStartPoint = lastRows + 2
               
             
            workbook.save(outputFile)
            
            # summary pack sale
            lastRows = self.lastRowNumbers(sheet)
            summaryPackSaleStartPoint = lastRows + 1
            
            for index, name in enumerate(offerName):
                summaryPackSaleStartPoint += 1
                sheet[f'A{summaryPackSaleStartPoint}'] = name
            
            workbook.save(outputFile)
            
            # team 1 agent 
            lastRows = self.lastRowNumbers(sheet)
            team1AgentStartPoint = lastRows + 1
            
            for index,name in enumerate(team1Label):
                team1AgentStartPoint += 1
                sheet[f'A{team1AgentStartPoint}'] = name
            
            
            workbook.save(outputFile)
            
            # team 1 pack sale
            lastRows = self.lastRowNumbers(sheet)
            team1PackSaleStartPoint = lastRows + 1
            
            for index,name in enumerate(offerName):
                # print(f'A{team1PackSaleStartPoint}')
                team1PackSaleStartPoint += 1
                sheet[f'A{team1PackSaleStartPoint}'] = name
            
             
            workbook.save(outputFile)
            
            # team 2 agent
            lastRows = self.lastRowNumbers(sheet)
            team2AgentStartPoint = lastRows + 1
            
            for index,name in enumerate(team2Label):
                team2AgentStartPoint += 1
                sheet[f'A{team2AgentStartPoint}'] = name
            
            
            workbook.save(outputFile)
            
            # team 2 pack sale 
            lastRows = self.lastRowNumbers(sheet)
            team2PackSaleStartPoint = lastRows + 1
            
            for index,name in enumerate(offerName):
                team2PackSaleStartPoint += 1
                sheet[f'A{team2PackSaleStartPoint}'] = name
            
            
            workbook.save(outputFile)
            
            print("Saved.", outputFile)
        except Exception as e:
            print(f"Error message: {e}")       