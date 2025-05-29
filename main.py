import pandas as pd
from openpyxl import load_workbook, Workbook
from write_label_dashboard import CreateLabel
from generate_dashboard import generate_dashboard_main
from total_summary_dashboard import total_summary_main
from file_fetch_from_ftp import ftp_file_fatch_main
from send_email import send_eamil
import datetime
import os
import shutil
from config import INPUT_BASE_DIR, OUTPUT_BASE_DIR

class main_app:
    def __init__(self):
        self.monthFormat = datetime.datetime(2025, 5, 10).strftime("%b")
        self.day = datetime.datetime(2025, 5, 10).strftime("%d")
        self.outputFile = f"{OUTPUT_BASE_DIR}\\Daily_Report_{self.monthFormat}.xlsx"
        self.productOfferFile = f"{INPUT_BASE_DIR}\\Product_Offer\\ProductOffer.xlsx"
        self.productOfferFileOld =  f"{INPUT_BASE_DIR}\\Product_Offer\\ProductOfferOld.xlsx"
        self.datetime_vrd = datetime.datetime(2025, 5, 10).strftime("%d-%b-%Y")
        self.datetime_mis = datetime.datetime(2025, 5, 10).strftime("%d%m%Y")
        self.workbook = load_workbook(self.outputFile)
        
    def _close_workbook(self):
        self.workbook.close()

    def run(self):
        try:
            # workbook = load_workbook(self.outputFile)
            sheet = self.workbook.active
            print("Excel file found.")
            print("Writing dashboard label...")
            
            if (sheet.max_column > 2):
                self._clear_total_column(sheet, sheet.max_column)
                
                self.workbook.save(self.outputFile)
                print("Clearing first column of cell...")
                
            self._write_dashboard_label()
        except FileNotFoundError:
            print("Excel file not found.")
            print("Creating File...")
            
            self.workbook = Workbook()
            
            self.workbook.save(self.outputFile)
            print("New workbook created.")
            
            self._write_dashboard_label()
            print("VRD Operation completed.")
    
    def _clear_total_column(self, sheet, max_col):
        sheet.delete_cols(max_col)

    def _check_product_offer_update(self, newOffer, oldOffer):
        # dropnan => NaN ignored
        # reset_index(drop=True) = index ignore 
        newList = newOffer['Offer Name'].dropna().reset_index(drop=True)
        oldList = oldOffer['Offer Name'].dropna().reset_index(drop=True)

        if len(newList) != len(oldList):
            return 'changed'
        elif not newList.equals(oldList):
            return 'changed'
        else:
            return 'no changed'
        
    def _update_old_product_offer(self):
        workbookNew = load_workbook(self.productOfferFile)
        workbookOld = load_workbook(self.productOfferFileOld)
        
        sheetNew = workbookNew.active
        sheetOld = workbookOld.active
        
        # clear all cell 
        for row in sheetOld.iter_rows(min_row=1, min_col=1):
            cell_coordinate_old = [cell.coordinate for cell in row]
            for index, cell_coordinate in enumerate(cell_coordinate_old):
                sheetOld[cell_coordinate] = ''
        
        workbookOld.save(self.productOfferFileOld)
        
        print("All cell cleared.")
        
        for row in sheetNew.iter_rows(min_row=1, min_col=1):
            cell_value = [cell.value for cell in row]
            cell_coordinate_new = [cell.coordinate for cell in row]
            for index, cell_coordinate in enumerate(cell_coordinate_new):
                sheetOld[cell_coordinate] = cell_value[index]
        
        workbookOld.save(self.productOfferFileOld)
        
        df = pd.read_excel(self.productOfferFile)
        df.to_csv(f"{INPUT_BASE_DIR}\\Product_Offer\\ProductOffer.csv", index=False, encoding="utf-8-sig")
        print("Product offer old updated.")
        
    def _clear_cell_of_label(self):
        sheet = self.workbook.active
        
        for row in range(1, sheet.max_row + 1):
            sheet[f"A{row}"].value = None
            
        
        self.workbook.save(self.outputFile)
        print("First columns of cell are cleared")

    def _cleaning_data(self):
        vmdFolderPath = f"{INPUT_BASE_DIR}\\VRD_Files\\{self.datetime_vrd}"
        
        try:
            for index, filename in enumerate(os.listdir(vmdFolderPath)):
                filePath = os.path.join(vmdFolderPath, filename)
                if os.path.isfile(filePath):
                    df = pd.read_csv(filePath)
                    df['AgentCID'] = df['AgentCID'].astype(str).str[:6]
                    df_sorted = df.sort_values(by=["OutboundNumber"], ascending=False)
                    df_sorted = df.sort_values(by=["AgentCID"], ascending=False)
                    df_sorted = df.sort_values(by=["TalkTime"], ascending=False)
                    
                    df_unique = df_sorted.drop_duplicates(subset=["AgentCID", "OutboundNumber"])
                    
                    formattedDirName = f"{INPUT_BASE_DIR}\\Formatted_VMD\\{self.datetime_vrd}"           
                    isExists = os.path.exists(formattedDirName)       
                    if not isExists:
                        os.mkdir(formattedDirName)         
                        
                    formattedVMDFilePath = os.path.join(formattedDirName, f"OUTBOUND_REPORT_VMD{index + 1}")
                    
                    df_unique.to_csv(formattedVMDFilePath+".csv", index=False, encoding="utf-8-sig")
        except FileNotFoundError:
            print(f"Folder not found at path {vmdFolderPath}")
        except Exception as e:
            print(f"Error message: {e}")
            
    def _label(self):
        label = {
                    'summary' : [
                        'Summary', 'Activities', 'Total Attempts Calls', 'Total Success Call', '% Of Total Vs Success Call', 'Total Login Agent', '% Success Call Vs Pack Sales', 'Total Revenue'
                    ],
                    'team1' : [
                        'Team 1', 'Activities', 'Total Attempts Calls', 'Total Success Call', '% Of Total Vs Success Call', 'Total Login Agent', '% Success Call Vs Pack Sales', 'Total Revenue'
                    ],
                    'team2' : [
                        'Team 2', 'Activities', 'Total Attempts Calls', 'Total Success Call', '% Of Total Vs Success Call', 'Total Login Agent', '% Success Call Vs Pack Sales', 'Total Revenue'
                    ]
            }
        
        return label 

    def _write_dashboard_label(self):
        offerName = pd.read_excel(self.productOfferFile, usecols=['Offer Name'])
        offerNameOld = pd.read_excel(self.productOfferFileOld, usecols=['Offer Name'])
        
        productOfferStatus = self._check_product_offer_update(offerName, offerNameOld)
        
        if (productOfferStatus == 'changed' or self.day == '01'):
            self._update_old_product_offer()
            self._clear_cell_of_label()
            print("Product offer changed, updating old product offer and clearing label cell.")
            agentInformationLabel = pd.DataFrame(self._label())
            
            labelCreation = CreateLabel(
                self.outputFile, agentInformationLabel['summary'], agentInformationLabel['team1'], agentInformationLabel['team2'], offerName
            )
            
            labelCreation.writeLabel()
        
    def _purging_dir(self):
        try:
            formattedVMDPath = f"{INPUT_BASE_DIR}\\Formatted_VMD"
            vmdFolderPath = f"{INPUT_BASE_DIR}\\VRD_Files"
            misFolderPath = f"{INPUT_BASE_DIR}\\MIS_Pack_Sale"

            for dir_name in os.listdir(vmdFolderPath):
                full_path = os.path.join(vmdFolderPath, dir_name)
                
                if os.path.exists(full_path) and os.path.isdir(full_path):
                    shutil.rmtree(full_path)
                else:
                    print("Folder not exists")
                    
            print(f"Purging completed at {vmdFolderPath}")
            
            for dir_name in os.listdir(misFolderPath):
                full_path = os.path.join(misFolderPath, dir_name)
                
                if os.path.exists(full_path) and os.path.isdir(full_path):
                    shutil.rmtree(full_path)
                else:
                    print("Folder not exists")
                    
            print(f"Purging completed at {misFolderPath}")
            
            for dir_name in os.listdir(formattedVMDPath):
                full_path = os.path.join(formattedVMDPath, dir_name)
                
                if os.path.exists(full_path) and os.path.isdir(full_path):
                    shutil.rmtree(full_path)
                else:
                    print("Folder not exists")
                    
            print(f"Purging completed at {formattedVMDPath}")
        except FileNotFoundError:
            print(f"Folder not found at path {vmdFolderPath}")
        except Exception as e:
            print(f"Error message: {e}")
            
def generate_dashboard():
    generate_dashboard_main()
    
def generate_total_summary_dashboard():
    total_summary_main()
    
def ftp_file_fetcher():
    ftp_file_fatch_main()
    
def send_email_main():
    app = send_eamil()
    app._email_sent_func()
            
__name__ = "__main__"
if __name__ == "__main__":
    app = main_app()
    ftp_file_fetcher()
    app._cleaning_data()
    app.run()
    generate_dashboard()
    print("Dashboard generation completed successfully.")
    generate_total_summary_dashboard()
    print("Total summary dashboard generation completed successfully.")
    app._close_workbook()
    print("Workbook closed.")
    print("VRD Operation completed successfully.")
    app._purging_dir()
    send_email_main()