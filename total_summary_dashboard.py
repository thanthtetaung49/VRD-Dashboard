import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from config import OUTPUT_BASE_DIR

class TotalSummary:
    def __init__(self):
        self.monthFormat = datetime.datetime(2025, 5, 10).strftime("%b")
        self.outputFile = f"{OUTPUT_BASE_DIR}\\Daily_Report_{self.monthFormat}.xlsx"
        self.datetime = datetime.datetime(2025, 5, 10).strftime("%y-%b-%d")
    
    def generate_summary(self):
        workbook = load_workbook(self.outputFile)
        sheet = workbook.active
        actual_max_col = sheet.max_column
        max_col = sheet.max_column + 1
        col_letter = get_column_letter(max_col)
        
        df = self._extract_sheet_data(sheet)
        
        self._write_formula_summary_total(sheet, df, col_letter)
        self._write_formula_total_success_percent(sheet, df, col_letter)
        self._write_formula_success_pack_percent(sheet, df, col_letter)
        self._write_formula_total_login_agent(sheet, df, col_letter, actual_max_col)
        self._set_headers(sheet, df, col_letter, actual_max_col)
        self._set_borders(sheet, max_col)
        self.autofit_columns(sheet)
        
        
        workbook.save(self.outputFile)
        
    def autofit_columns(self, sheet):
        for col in sheet.columns:
            column = col[0].column  
            column_letter = get_column_letter(column)  
            sheet.column_dimensions[column_letter].width = 10
            sheet.column_dimensions['A'].width = 25
        
    def _set_borders(self, sheet, max_col):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        alignment = Alignment(horizontal='right')
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=max_col):
            for cell in row:
                cell.alignment = alignment
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
        
    def _set_headers(self, sheet, df, col_letter, actual_max_col):   
        filtered_df = df[df['cell_value'].notna()].copy()
        filtered_df = df[
            df['cell_value'].isin(['Activities', 'Pack Sale'])
        ]['cell_coordinate'].to_list()
        
        coord_numbers = [coord[1:] for coord in filtered_df]
        cell_coord_1 = [f"{col_letter}{coord}" for coord in coord_numbers]
        
        for cell in cell_coord_1:
            # print(cell)
            sheet[cell] = 'Total'
            
            fill = PatternFill(start_color='a5d1f2', end_color='a5d1f2', fill_type='solid')
            bold_font = Font(bold=True)
            sheet.freeze_panes = None
            
            sheet[cell].fill = fill
            sheet[cell].font = bold_font
                
        
    def _extract_sheet_data(self, sheet):
        data = []
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                data.append({'cell_coordinate': cell.coordinate, 'cell_value': cell.value})
        return pd.DataFrame(data)
    
    def _write_formula_summary_total(self, sheet, df, col_letter):
        filtered_df = df[df['cell_value'].notna()].copy()
        filtered_df = df[
            ~df['cell_coordinate'].str.startswith('A') &
            df['cell_value'].notna() &
            ~df['cell_value'].isin([self.datetime, '% Of Total Vs Success Call', '% Success Call Vs Pack Sales', 'Total Login Agent', 'Activities', 'Pack Sale'])
        ]
        
        filtered_df = filtered_df.copy()
        filtered_df['cell_coord_number'] = filtered_df['cell_coordinate'].astype(str).str[1:]
        
        grouped = filtered_df.groupby('cell_coord_number')['cell_coordinate'].apply(list).reset_index()
        
        grouped['writable_coord'] = grouped['cell_coord_number'].apply(lambda x: col_letter + x)
        
        for _, row in grouped.iterrows():
            writable_coord = row['writable_coord']
            total_formula = f"=SUM({','.join(row['cell_coordinate'])})"
            # print(total_formula)
            
            sheet[writable_coord] = total_formula
    
    def _write_formula_total_success_percent(self, sheet, df, col_letter):
        filtered_total_success_coord = df[df['cell_value'].isin(['% Of Total Vs Success Call'])].copy()['cell_coordinate'].to_list()
        
        total_attempsts_call = df[df['cell_value'].isin(['Total Attempts Calls'])].copy()['cell_coordinate'].to_list()
        total_success_call = df[df['cell_value'].isin(['Total Success Call'])].copy()['cell_coordinate'].to_list()
        
        mapping_cell_coords = [cell_coord for cell_coord in zip(filtered_total_success_coord, total_attempsts_call, total_success_call)]
        
        for cell in mapping_cell_coords:
            cell_coord_1 = f'{col_letter}{cell[0][1:]}'
            cell_coord_2 = f'{col_letter}{cell[1][1:]}'
            cell_coord_3 = f'{col_letter}{cell[2][1:]}'
            
            sheet[cell_coord_1] = f'=TEXT(ROUND({cell_coord_3}/{cell_coord_2}*100, 0), "0") & "%"'
    
    def _write_formula_success_pack_percent(self, sheet, df, col_letter):
        filtered_success_pack_coord = df[df['cell_value'].isin(['% Success Call Vs Pack Sales'])].copy()['cell_coordinate'].to_list()
        
        total_success_call = df[df['cell_value'].isin(['Total Success Call'])].copy()['cell_coordinate'].to_list()
        total_pack = df[df['cell_value'].isin(['Total'])].copy()['cell_coordinate'].to_list()
        
        mapping_cell_coords = [cell_coord for cell_coord in zip(filtered_success_pack_coord, total_success_call, total_pack)]
        
        for cell in mapping_cell_coords:
            cell_coord_1 = f'{col_letter}{cell[0][1:]}'
            cell_coord_2 = f'{col_letter}{cell[1][1:]}'
            cell_coord_3 = f'{col_letter}{cell[2][1:]}'
            
            sheet[cell_coord_1] = f'=TEXT(ROUND({cell_coord_3}/{cell_coord_2}*100, 0), "0") & "%"'
            
    def _write_formula_total_login_agent(self, sheet, df, col_letter, actual_max_col):
        filtered_df = df[df['cell_value'].notna()].copy()
        
        cell_coord = df[
            df['cell_value'].isin(['Total Login Agent'])
        ]['cell_coordinate'].to_list()
        
        coord_numbers = [coord[1:] for coord in cell_coord]
        
        cell_coord_1 = [f"{col_letter}{coord}" for coord in coord_numbers]
        
        start_column = 2
        end_column = actual_max_col
        
        all_coords = []
        
        for i in range(start_column, end_column + 1):
            col_letter = get_column_letter(i)
            cell_coord = [f"{col_letter}{coord}" for coord in coord_numbers]
            
            all_coords.append(cell_coord)
            
        for cell in zip(cell_coord_1, *all_coords):
            formula = f"=ROUNDUP(AVERAGE({','.join(cell[1:])}))"
            sheet[cell[0]] = formula
            
def total_summary_main():
    total_summary = TotalSummary()
    total_summary.generate_summary()
    print("Total summary generated successfully.")
    
if __name__ == "__main__":
    total_summary_main()